from openpyxl import load_workbook  # Excel reader
import confparser  # Author: Tim Dorssers
from jinja2 import Template  # Templates
import logging.config  # Logger
import time
import warnings  # Used to supress warnings from Excel reader
from collections import ChainMap  # used to merge multiple dictionaries into 1
from datetime import datetime  # For timing execution time & output file naming
import os


def main(core_switch, rack_port_input, excluded_svi_input, svi_nac):
    used_interfaces = []
    counted_stacks = []
    today = datetime.now()
    capacity = False
    # Read the input xl, retrieve the matrix & to be port allocations
    old_inventory, inventory_matrix = port_availability_workbook_reader(wb_input=rack_port_input)
    # Read core switch to find the active SVIs
    nac_svi = []
    if svi_nac:
        nac_svi = svi_discovery(core_switch, excluded_svi_input)  # TODO: SVI logic
    # Grab the config of the old switch, parse it into a variable
    for key in inventory_matrix:  # Start looping through switches
        old_config = config_parse(key + '.log')
        templated_config = []
        old_interfaces = []
        # Get the interface changes from the output of the xl
        for old_switch_stack in old_inventory:
            if key == old_switch_stack['old_sw_name']:
                old_interfaces.append(old_switch_stack['interfaces'])  # FROM THE EXCEL FILE
                capacity = old_switch_stack['capacity']
        interface_matrix = dict(ChainMap(*old_interfaces))  # merge the entire stack of interfaces into 1 dictionary
        # Build the templates according to the old interface names
        # Interface Matrix:
        # At this point we have a matrix where the key is the existing interface, and the value is the destination
        for used_interface in old_config["interface"]:  # For every interface in the old config file.
            logger.debug(used_interface + " was returned when looping interfaces in the old config file")
            # I want to take the old interface, and match it against the destined interface
            if isinstance(old_config["interface"][used_interface]['mode'], list):
                logger.error("Show Running Config is duplicated in the file: " + key + ".log!")
                exit()
            if old_config["interface"][used_interface]["mode"] == "access":  # but only if it's an access interface
                logger.debug(used_interface + " is an access interface")
                int_refactor = used_interface.split('net')[1].split(
                    '/')  # Start to format the interface to be the same as interface_matrix
                if get_number_of_elements(int_refactor) == 3:  # fix incase we encounter FA interfaces
                    int_refactor.pop(1)  # fix incase we encounter FA interfaces
                int_match = int_refactor[0] + '/' + int_refactor[1]  # Formatting done.
                target_interface = interface_matrix[int_match]  # Find the target interface

                if target_interface == "F":  # Don't do anything for unused ports just yet.
                    continue
                else:
                    target_interface = target_interface.split("/")
                    target_interface = "GigabitEthernet" + target_interface[0] + "/0/" + target_interface[1]  # format
                logger.info(used_interface + " moves to " + target_interface)
                # Prepare variables for input
                description = False
                port_security = False
                dhcp_snooping = False
                interface_aaa = "nonac"  # Default as NO NAC Interface
                voice_vlan = False
                ps_max = False
                ps_aging_time = False
                access_vlan = old_config['interface'][used_interface]["access_vlan"]  # Retain previous data VLAN
                if old_config['interface'][used_interface]["description"]:
                    description = old_config['interface'][used_interface]["description"]  # Retain old value
                if old_config['interface'][used_interface]["voice_vlan"]:
                    voice_vlan = old_config['interface'][used_interface]["voice_vlan"]  # Retain old value
                if old_config['interface'][used_interface]["ps_max"]:
                    ps_max = old_config['interface'][used_interface]["ps_max"]  # Retain old value
                    port_security = True
                if old_config['interface'][used_interface]["ps_aging_time"]:
                    ps_aging_time = old_config['interface'][used_interface]["ps_aging_time"]  # Retain old value
                if old_config['interface'][used_interface]["dhcp_snooping"]:
                    dhcp_snooping = True  # Retain old value
                if old_config['interface'][used_interface]["dot1x_pae"] == "authenticator":
                    interface_aaa = "nac_enforce"  # If dot1x pae is configured, default to NAC enforce mode
                if old_config['interface'][used_interface]["auth_open"]:
                    interface_aaa = "nac_open"  # If authentication open is configured, default to NAC Open mode
                if svi_nac:  # If this feature is enabled, it will override anything discovered on the old interface
                    # for NAC.  It helps correct potential mistakes in old config.
                    if access_vlan in nac_svi:
                        interface_aaa = "nac_open"
                    if access_vlan not in nac_svi:
                        interface_aaa = "nonac"

                input_list = {
                    "interface": target_interface,
                    "data_vlan": access_vlan,
                    "description": description,
                    "voice_vlan": voice_vlan,
                    "ps_max": ps_max,
                    "port_security_age": ps_aging_time,
                    "dhcp_snooping": dhcp_snooping,
                    "port_security": port_security
                }
                interface_config = templater(input_list, interface_aaa)  # Build interface config

                templated_config.append(interface_config)  # Append interface config to list
                used_interfaces.append(target_interface.split('net')[1])
                for counted_interface in used_interfaces:
                    if counted_interface.split('/')[0] not in counted_stacks:
                        counted_stacks.append(counted_interface.split('/')[0])

        logger.debug(templated_config)
        # Discover the uplink detail for the switch inside the loop
        uplinks = core_port_workbook_reader(wb_input=rack_port_input, cp_wb_input=inventory_matrix[key])
        for uplink_interface in uplinks:
            input_list = {  # Build the variables for J2 template
                "interface": uplink_interface['access_side_interface'],
                "description": "To " + uplink_interface['core_hostname'] + " " + uplink_interface['core_side_interface']
            }
            trunk_config = templater(input_list, 'up_trunk')
            templated_config.append(trunk_config)
        input_list = {
            "description": "To " + uplinks[0]['core_hostname']
        }
        port_channel_config = templater(input_list, 'etherchannel')
        templated_config.append(port_channel_config)  # Add to the output to be written

        # TODO: Misc Trunks go here

        # Calculate unused interfaces
        logger.debug(capacity)
        logger.debug(used_interfaces)
        all_interfaces = []
        with open("logic_templates/all_interfaces_48.j2") as f:
            template_file = f.read()
        t = Template(template_file)
        for stack in counted_stacks:
            # Quick way to render total amount of interfaces available.  Based on the total number of stacks being
            # configured.  IMPORTANT: If a stack is being empty and is empty according to port allocation, those ports
            # will not be shutdown through this process.  Script only knows of interfaces and stacks existing to the
            # left of the document.
            stack_all_interfaces = t.render({"stack": stack}).split('\n')
            all_interfaces.append(stack_all_interfaces)
        all_interfaces = [j for i in all_interfaces for j in i]  # merge lists into a single list
        # discover unused interfaces by subtracting used interfaces from all interfaces
        free_interfaces = [x for x in all_interfaces if x not in used_interfaces or used_interfaces.remove(x)]
        for free_interface in free_interfaces:
            shutdown_interface = templater({"interface": "GigabitEthernet" + free_interface}, 'unused')
            templated_config.append(shutdown_interface)
        # Write the output for the switch.
        with open(inventory_matrix[key] + '-' + today.strftime("%d%b%Y%-H%M%S") + '.txt', 'a') as file:
            for command in templated_config:
                file.write(command)
            file.close()


def templater(input_lst, interface_type):
    with open("configs/interface_" + interface_type + ".j2", "r") as interface_file:
        interface_template = interface_file.read()
    t = Template(interface_template)  # loads template
    configuration = t.render(input_lst)  # renders template
    return configuration


def svi_discovery(core_sw, excluded_svi_list):
    output = []
    core_config = config_parse(core_sw)
    for interface in core_config['interface']:
        if 'Vlan' in interface and 'shutdown' not in core_config['interface'][interface]:
            subnet_mask = False
            if core_config['interface'][interface]['ipv4']:
                subnet_mask = int(core_config['interface'][interface]['ipv4'].split('/')[1])
            else:
                logger.warning(interface + ' has no IP address configured on this active svi, ignoring this SVI')
            if subnet_mask and subnet_mask < 28:
                target_interface = interface.split("n", 1)
                output.append(target_interface[1])
    output = list(set(output) - set(excluded_svi_list))  # Remove any VLANs specifically set to exclude
    return output


def config_parse(file):
    if not os.path.exists(file):
        logger.critical(file + ": This file does not exist in the directory!")
        exit()
    dissector = confparser.Dissector.from_file('ios.yaml')
    return dissector.parse_file(file)


def core_port_workbook_reader(wb_input, cp_wb_input):
    wb = load_workbook(filename=wb_input, data_only=True)
    sheet_name = wb.sheetnames[index_containing_substring(wb.sheetnames, "Core Port Layout")]
    sheet = wb[sheet_name]
    e_cell = 1
    output = []
    while e_cell < 10000:
        if sheet["E" + str(e_cell)].value == cp_wb_input:
            output.append({"core_hostname": sheet["A" + str(e_cell)].value,
                           "core_side_interface": sheet["B" + str(e_cell)].value,
                           "access_hostname": sheet["E" + str(e_cell)].value,
                           "access_side_interface": sheet["D" + str(e_cell)].value})
            e_cell += 1
        else:
            e_cell += 1
    return output


def port_availability_workbook_reader(wb_input):
    wb = load_workbook(filename=wb_input, data_only=True)
    sheet_name = wb.sheetnames[index_containing_substring(wb.sheetnames, "availability")]
    sheet = wb[sheet_name]
    e_cell = 1
    old_inventory = []
    inventory_matrix = {}
    while e_cell < 10000:
        output = {"interfaces": {}}
        if sheet["E" + str(e_cell)].value == "Switch Name: ":
            old_sw = sheet["H" + str(e_cell)].value
            new_sw = sheet["Y" + str(e_cell)].value
            stack_member = sheet["R" + str(e_cell)].value
            output.update({"old_stack_id": stack_member})
            output.update({"old_sw_name": old_sw})
            inventory_matrix.update({old_sw: new_sw})
            switch_capacity = 0
            if sheet["D" + str(e_cell + 1)].value == 1:
                switch_capacity = 48
            elif sheet["P" + str(e_cell + 1)].value == 1:
                switch_capacity = 24
            output.update({"capacity": switch_capacity})
            port = "D"
            port_number = 1
            row = 1
            result = False
            while port_number < switch_capacity + 1:
                if "Z" not in port or "AA" not in port:
                    if not sheet[port + str(e_cell + 2)].value:
                        result = str(stack_member) + "/" + str(port_number)
                    if sheet[port + str(e_cell + 2)].value:
                        result = sheet[port + str(e_cell + 2)].value

                    output['interfaces'].update({
                        str(stack_member) + "/" + str(port_number): result
                    })
                    port_number += 2
                if port == "AA":
                    if port_number < switch_capacity + 1:
                        if not sheet[port + str(e_cell + 2)].value:
                            result = str(stack_member) + "/" + str(port_number)
                        if sheet[port + str(e_cell + 2)].value:
                            result = sheet[port + str(e_cell + 2)].value

                        output['interfaces'].update({
                            str(stack_member) + "/" + str(port_number): result
                        })
                    e_cell += 1
                    port = "D"
                    port_number = 2
                    if row == 1:
                        row = 2
                    elif row == 2:
                        port_number = 100
                else:
                    port = chr(ord(port) + 1)
                if port == "Z":
                    if not sheet[port + str(e_cell + 2)].value:
                        result = str(stack_member) + "/" + str(port_number)
                    if sheet[port + str(e_cell + 2)].value:
                        result = sheet[port + str(e_cell + 2)].value

                    output['interfaces'].update({
                        str(stack_member) + "/" + str(port_number): result
                    })
                    port_number += 2
                    port = "AA"

            e_cell += 1
            old_inventory.append(output)
        else:
            e_cell += 1

    return old_inventory, inventory_matrix


def index_containing_substring(the_list, substring):
    for i, s in enumerate(the_list):
        if substring in s:
            return i
    return -1


def get_number_of_elements(input_list):
    count = 0
    for element in input_list:
        logger.debug(element)
        count += 1
    return count


if __name__ == "__main__":
    start_time = time.time()
    # setup logging
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    logging.config.fileConfig(fname='logging.conf', disable_existing_loggers=True)
    stream = logging.StreamHandler()
    stream_format = logging.Formatter("%(levelname)s:%(module)s:%(lineno)d:%(message)s")
    stream.setFormatter(stream_format)
    logger = logging.getLogger(__name__)
    logger.propagate = False
    logger.addHandler(stream)

    core_file = "CSW.log"
    rack_port_xl = "input.xlsx"

    # The following 2 variables toggle and configure the NAC SVI configuration.  If you want to override the old
    # switch configuration and apply NAC to a port based on if the associated SVI exists on the core switch, set this
    # to True (it is by default).  If you want to use this feature, but exclude specific VLANs, add them to the
    # excluded_svi_lst as integers.
    SVI_NAC_option = True
    excluded_svi_lst = []

    main(core_file, rack_port_xl, excluded_svi_lst, SVI_NAC_option)
    run_time = time.time() - start_time
    run_time = round(run_time, 2)
    logger.info("Time to run: " + str(run_time) + " seconds")
